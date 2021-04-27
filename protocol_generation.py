from typing import List, Optional, Tuple, Dict
import logging
import odf
import abc
import os

import openpyxl

from upms_measure import UpmsMeasure
from text import Text


class UpmsProtocolGenerator(abc.ABC):
    DATA_SHEET_RU = "Данные"
    PHOTO_SHEET_RU = "Фото"
    DATA_SHEET_EN = "Data"
    PHOTO_SHEET_EN = "Photos"

    @abc.abstractmethod
    def is_template_ok(self):
        pass

    @staticmethod
    @abc.abstractmethod
    def protocol_ext() -> str:
        pass

    @abc.abstractmethod
    def get_report_path(self) -> str:
        pass

    @abc.abstractmethod
    def insert_measures(self, a_upms_measures: List[UpmsMeasure]):
        pass

    @abc.abstractmethod
    def insert_photos(self, a_photos_path, a_upms_measures: List[UpmsMeasure]):
        pass

    @abc.abstractmethod
    def insert_extra_parameters(self, a_extra_parameters: List[Tuple[str, str]]):
        pass

    @abc.abstractmethod
    def save(self):
        pass


class ExcelProtocolGenerator(UpmsProtocolGenerator):
    def __init__(self, a_report_filepath: str):
        self.report_filepath = a_report_filepath

        self.template_is_ok = True
        self.workbook = openpyxl.load_workbook(self.report_filepath)

        if self.DATA_SHEET_RU in self.workbook.sheetnames:
            self.data_sheet = self.DATA_SHEET_RU
            self.photo_sheet = self.PHOTO_SHEET_RU
        elif self.DATA_SHEET_EN in self.workbook.sheetnames:
            self.data_sheet = self.DATA_SHEET_EN
            self.photo_sheet = self.PHOTO_SHEET_EN
        else:
            self.data_sheet = ""
            self.photo_sheet = ""
            self.template_is_ok = False

    def __del__(self):
        self.workbook.close()

    def save(self):
        self.workbook.save(self.report_filepath)

    def is_template_ok(self):
        return self.template_is_ok

    @staticmethod
    def protocol_ext() -> str:
        return ".xlsx"

    def get_report_path(self) -> str:
        return self.report_filepath

    def insert_measures(self, a_upms_measures: List[UpmsMeasure]):
        if self.template_is_ok:
            sheet = self.workbook.get_sheet_by_name(self.data_sheet)
            for idx, column in enumerate(sheet.iter_cols(min_col=2, max_col=len(a_upms_measures) + 1, min_row=1, max_row=5)):
                column[0].value = a_upms_measures[idx].id
                column[1].value = a_upms_measures[idx].date
                column[2].value = a_upms_measures[idx].interval
                column[3].value = a_upms_measures[idx].result
                column[4].value = a_upms_measures[idx].comment

    def insert_photos(self, a_photos_path: str, a_upms_measures: List[UpmsMeasure]):
        if self.template_is_ok:
            if self.photo_sheet in self.workbook.sheetnames:
                sheet = self.workbook.get_sheet_by_name(self.photo_sheet)
            else:
                sheet = self.workbook.create_sheet(self.photo_sheet)

            vertical_start = 3
            vertical_step = 27

            horizontal_start = 2
            horizontal_step = 9
            for idx, upms_measure in enumerate(a_upms_measures):
                photo_path = a_photos_path.rstrip(os.sep) + os.sep + f"{upms_measure.id}.jpg"
                try:
                    img = openpyxl.drawing.image.Image(photo_path)
                    new_height = img.height * 2 / 3
                    new_width = img.width * 2 / 3
                    img.height = new_height
                    img.width = new_width

                    row = vertical_start + idx // 2 * vertical_step if idx % 2 == 0 else \
                        vertical_start + idx // 2 * vertical_step + 1
                    col = horizontal_start if idx % 2 == 0 else \
                        (horizontal_start + horizontal_step)

                    sheet.add_image(img, sheet.cell(row, col).coordinate)
                    cell_font = openpyxl.styles.Font(size='15')
                    sheet.cell(row - 1, col, value=Text.get("measure").format(upms_measure.id)).font = cell_font
                except FileNotFoundError:
                    logging.warning(f"File {photo_path} is not found")

    def insert_extra_parameters(self, a_extra_parameters: List[Tuple[str, str]]):
        if self.template_is_ok:
            sheet = self.workbook.get_sheet_by_name(self.data_sheet)
            row = 8
            for parameter, value in a_extra_parameters:
                sheet.cell(row, 1).value = parameter
                sheet.cell(row, 2).value = value
                sheet.row_dimensions[row].height = 30
                row += 1


class CalcProtocolGenerator(UpmsProtocolGenerator):
    def __init__(self, a_report_filepath: str):
        self.report_filepath = a_report_filepath

        self.template_is_ok = True
        self.workbook = None

        if self.DATA_SHEET_RU in self.workbook.sheetnames:
            self.data_sheet = self.DATA_SHEET_RU
            self.photo_sheet = self.PHOTO_SHEET_RU
        elif self.DATA_SHEET_EN in self.workbook.sheetnames:
            self.data_sheet = self.DATA_SHEET_EN
            self.photo_sheet = self.PHOTO_SHEET_EN
        else:
            self.data_sheet = ""
            self.photo_sheet = ""
            self.template_is_ok = False

    def __del__(self):
        pass

    def is_template_ok(self):
        return self.template_is_ok

    @staticmethod
    def protocol_ext() -> str:
        return ".ods"

    def get_report_path(self) -> str:
        return self.report_filepath

    def insert_measures(self, a_upms_measures: List[UpmsMeasure]):
        pass

    def insert_photos(self, a_photos_path, a_upms_measures: List[UpmsMeasure]):
        pass

    def insert_extra_parameters(self, a_extra_parameters: List[Tuple[str, str]]):
        pass

    def save(self):
        pass
