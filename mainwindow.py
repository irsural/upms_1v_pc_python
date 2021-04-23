from logging.handlers import RotatingFileHandler
from typing import List, Optional, Tuple, Dict
import logging
import shutil
import socket
import sys
import os

from PyQt5 import QtWidgets, QtCore, QtGui, QtNetwork
import openpyxl

from irspy.qt.custom_widgets.QTableDelegates import TransparentPainterForView
from irspy.utils import exception_decorator, exception_decorator_print
from irspy.settings_ini_parser import BadIniException
from irspy.qt import qt_utils

from extra_parameters_dialog import ExtraParametersDialog
from ui.py.mainwindow import Ui_MainWindow as MainForm
from result_input_dialog import ResultInputDialog
from upms_db_model import UpmsDatabaseModel
from upms_database import UpmsDatabase
from about_dialog import AboutDialog
from upms_measure import UpmsMeasure
from text import Text
import upms_tftp
import settings


ENGLISH_VERSION = True


class MainWindow(QtWidgets.QMainWindow):
    measures_filename = "main_table.csv"
    default_name_template = "report"

    MEASURE_TYPE_TO_TEMPLATE_PATH = {
        UpmsMeasure.MeasureType.MECH_STOPWATCH: ("./Templates/ms_template.xlsx", "./Templates/ms_template.ods"),
        UpmsMeasure.MeasureType.ELEC_STOPWATCH: ("./Templates/es_template.xlsx", "./Templates/es_template.ods"),
        UpmsMeasure.MeasureType.CLOCK: ("./Templates/clock_template.xlsx", "./Templates/clock_template.ods"),
    }

    data_sheet_ru = "Данные"
    photo_sheet_ru = "Фото"
    data_sheet_en = "Data"
    photo_sheet_en = "Photos"

    def __init__(self):
        super().__init__()

        self.ui = MainForm()
        self.ui.setupUi(self)

        try:
            self.settings = settings.get_ini_settings()
            ini_ok = True
        except BadIniException:
            ini_ok = False
            QtWidgets.QMessageBox.critical(self, Text.get("err"), Text.get("ini_err"))
        if ini_ok:
            self.settings.restore_qwidget_state(self)
            self.settings.restore_qwidget_state(self.ui.mw_splitter_1)
            self.settings.restore_qwidget_state(self.ui.measures_table)

            self.ui.download_progress_bar.setHidden(True)

            if ENGLISH_VERSION:
                self.ui.russian_language_action.setVisible(False)
                self.settings.language = Text.Lang.EN

            self.translator = QtCore.QTranslator(QtWidgets.QApplication.instance())
            self.install_translator(self.settings.language)

            if self.settings.language == Text.Lang.EN:
                self.ui.english_language_action.setChecked(True)
            else:
                self.ui.russian_language_action.setChecked(True)

            self.set_up_logger()

            self.ui.measures_table.setItemDelegate(TransparentPainterForView(self.ui.measures_table, "#d4d4ff"))
            self.ui.download_path_edit.setText(self.settings.save_folder_path)

            self.ui.ip_combobox.setEditText(self.settings.ip)
            self.ui.download_path_edit.setText(self.settings.path)
            self.ui.name_template_edit.setText(self.settings.name_template)
            self.ui.save_folder_edit.setText(self.settings.save_folder)

            self.db = UpmsDatabase("database.db")
            self.measures_table_model: Optional[UpmsDatabaseModel] = None
            self.proxy: Optional[QtCore.QSortFilterProxyModel] = None
            self.update_model()

            self.bcast_sockets: Dict[str, QtNetwork.QUdpSocket] = {}
            self.get_broadcast_ips()

            self.show()
            self.connect_all()

            self.ui.get_ip_button.click()
        else:
            self.close()

    def install_translator(self, a_language: Text.Lang):
        try:
            # Для программы, собранной pyinstaller
            base_path = sys._MEIPASS
        except AttributeError:
            base_path = "."

        translate_file = Text.LANG_TO_QT_TRANSLATE[a_language]
        self.translator.load(os.path.join(base_path, translate_file))
        QtWidgets.QApplication.instance().installTranslator(self.translator)
        self.ui.retranslateUi(self)
        Text.set_language(a_language)

    def connect_all(self):
        self.ui.open_about_action.triggered.connect(self.open_about)
        self.ui.download_from_upms_button.clicked.connect(self.download_from_upms_button_clicked)
        self.ui.ip_combobox.editTextChanged.connect(self.ip_changed)
        self.ui.download_path_edit.textChanged.connect(self.path_changed)
        self.ui.name_template_edit.textChanged.connect(self.name_template_changed)
        self.ui.save_folder_edit.textChanged.connect(self.save_folder_changed)

        self.ui.get_ip_button.clicked.connect(self.get_ip_button_clicked)

        self.ui.input_result_button.clicked.connect(self.input_result_button_clicked)
        self.ui.create_report_button.clicked.connect(self.create_report_button_clicked)
        self.ui.remove_selected_button.clicked.connect(self.remove_selected_button_clicked)

        self.ui.choose_download_path_button.clicked.connect(self.choose_download_path_button_clicked)
        self.ui.choose_save_folder_path_edit.clicked.connect(self.choose_save_folder_path_button_clicked)
        self.ui.extra_params_button.clicked.connect(self.extra_params_button_clicked)

        self.ui.russian_language_action.triggered.connect(self.russian_language_chosen)
        self.ui.english_language_action.triggered.connect(self.english_language_chosen)


        group = QtWidgets.QActionGroup(self)
        group.addAction(self.ui.russian_language_action)
        group.addAction(self.ui.english_language_action)

    def set_up_logger(self):
        log = qt_utils.QTextEditLogger(self.ui.log_text_edit)
        log.setFormatter(logging.Formatter('%(asctime)s - %(message)s', datefmt='%H:%M:%S'))

        file_log = RotatingFileHandler("upms_1v_pc.log", maxBytes=30*1024*1024, backupCount=3, encoding='utf8')
        file_log.setLevel(logging.DEBUG)
        file_log.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s', datefmt='%H:%M:%S'))

        logging.getLogger().addHandler(file_log)
        logging.getLogger().addHandler(log)
        logging.getLogger().setLevel(logging.WARNING)

    def update_model(self):
        self.measures_table_model = UpmsDatabaseModel(self.db, self)
        self.proxy = QtCore.QSortFilterProxyModel()
        self.proxy.setSourceModel(self.measures_table_model)
        self.ui.measures_table.setModel(self.proxy)
        self.ui.measures_table.resizeRowsToContents()

    def download_measures_list(self, a_download_filepath: str) -> bool:
        result = False
        ip = self.ui.ip_combobox.currentText()
        try:
            socket.inet_aton(ip)
        except socket.error:
            QtWidgets.QMessageBox.critical(self, Text.get("err"), Text.get("ip_format_err"),
                                           QtWidgets.QMessageBox.Ok, QtWidgets.QMessageBox.Ok)
        else:
            try:
                files_list = upms_tftp.get_files_list(self.ui.ip_combobox.currentText())
            except ConnectionResetError:
                QtWidgets.QMessageBox.critical(self, Text.get("err"), Text.get("connection_err"),
                                               QtWidgets.QMessageBox.Ok, QtWidgets.QMessageBox.Ok)
            else:
                if files_list is not None:
                    if self.measures_filename in files_list:
                        try:
                            upms_tftp.download_file_by_tftp(ip, self.measures_filename, a_download_filepath)
                            result = True
                        except FileNotFoundError:
                            QtWidgets.QMessageBox.critical(self, Text.get("err"), Text.get("path_err"),
                                                           QtWidgets.QMessageBox.Ok, QtWidgets.QMessageBox.Ok)
                    else:
                        QtWidgets.QMessageBox.critical(self, Text.get("err"),
                                                       Text.get("main_table_not_found_err").format(self.measures_filename),
                                                       QtWidgets.QMessageBox.Ok, QtWidgets.QMessageBox.Ok)
                else:
                    QtWidgets.QMessageBox.critical(self, Text.get("err"), Text.get("download_err"),
                                                   QtWidgets.QMessageBox.Ok, QtWidgets.QMessageBox.Ok)
        return result

    def download_photo(self, a_measure_id: int, a_files_list: List[str], a_download_folder: str):
        photo_name = f"{a_measure_id}.jpg"
        download_path = a_download_folder.rstrip(os.sep) + os.sep + photo_name
        if photo_name in a_files_list:
            logging.warning(f"Download {photo_name}")
            upms_tftp.download_file_by_tftp(self.ui.ip_combobox.currentText(), photo_name, download_path)

    @exception_decorator
    def download_from_upms_button_clicked(self, _):
        download_folder = self.ui.download_path_edit.text()
        if download_folder and os.path.isdir(download_folder):
            download_filepath = download_folder.rstrip(os.sep) + os.sep + self.measures_filename

            if self.download_measures_list(download_filepath):
                with open(download_filepath, 'r', encoding='utf8') as measures_list_file:
                    # Первая строка - заголовок
                    measures_list = [file for idx, file in enumerate(measures_list_file) if idx != 0]

                update_all = False
                upms_files_list = upms_tftp.get_files_list(self.ui.ip_combobox.currentText())
                self.ui.download_progress_bar.setHidden(False)
                for number, measure in enumerate(measures_list):
                    self.ui.download_progress_bar.setValue(
                        number / (len(upms_files_list) - 1) * self.ui.download_progress_bar.maximum())

                    measure_params = [meas.strip() for meas in measure.strip().split(',')]
                    upms_measure = UpmsMeasure(*measure_params)
                    upms_prev_measure = self.db.get(upms_measure.id)

                    if upms_measure != upms_prev_measure:
                        if upms_prev_measure is None:
                            self.db.add(upms_measure)
                            self.download_photo(upms_measure.id, upms_files_list, download_folder)
                        else:
                            if update_all:
                                res = QtWidgets.QMessageBox.Yes
                            else:
                                res = QtWidgets.QMessageBox.question(
                                    self, Text.get("warning"), Text.get("ask_overwrite").format(upms_measure.id),
                                    QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.YesAll |
                                    QtWidgets.QMessageBox.No | QtWidgets.QMessageBox.Cancel, QtWidgets.QMessageBox.Yes
                                )
                            if res == QtWidgets.QMessageBox.Yes:
                                self.db.update(upms_measure)
                                self.download_photo(upms_measure.id, upms_files_list, download_folder)
                            elif res == QtWidgets.QMessageBox.YesAll:
                                self.db.update(upms_measure)
                                self.download_photo(upms_measure.id, upms_files_list, download_folder)
                                update_all = True
                            elif res == QtWidgets.QMessageBox.Cancel:
                                QtWidgets.QMessageBox.warning(self, Text.get("warning"), Text.get("download_canceled"),
                                                              QtWidgets.QMessageBox.Ok, QtWidgets.QMessageBox.Ok)
                                break
                    elif not os.path.isfile(download_folder.rstrip(os.sep) + os.sep + f"{upms_measure.id}.jpg"):
                        # Докачиваем файл, если запись в БД есть, а файла нет
                        self.download_photo(upms_measure.id, upms_files_list, download_folder)
                else:
                    QtWidgets.QMessageBox.information(self, Text.get("info"), Text.get("success_download"),
                                                      QtWidgets.QMessageBox.Ok, QtWidgets.QMessageBox.Ok)

                self.ui.download_progress_bar.setValue(0)
                self.ui.download_progress_bar.setHidden(True)
                self.update_model()

        else:
            QtWidgets.QMessageBox.critical(self, Text.get("err"), Text.get("path_err"),
                                           QtWidgets.QMessageBox.Ok, QtWidgets.QMessageBox.Ok)

    def ip_changed(self, a_ip):
        self.settings.ip = a_ip

    def path_changed(self, a_path):
        self.settings.path = a_path

    def name_template_changed(self, a_name_template):
        self.settings.name_template = a_name_template

    def save_folder_changed(self, a_save_folder):
        self.settings.save_folder = a_save_folder

    def get_broadcast_ips(self):
        localhost = QtNetwork.QHostAddress(QtNetwork.QHostAddress.LocalHost)
        for _if in QtNetwork.QNetworkInterface.allInterfaces():
            for addr in _if.addressEntries():
                if addr.ip().protocol() == QtNetwork.QAbstractSocket.IPv4Protocol and not addr.netmask().isNull() and \
                        addr.ip() != localhost:
                    bcast = QtNetwork.QHostAddress(addr.ip().toIPv4Address() |
                                                   ((1 << 32) - 1 - addr.netmask().toIPv4Address()))  # bit not
                    udp_sock = QtNetwork.QUdpSocket(self)
                    udp_sock.bind(addr.ip())

                    self.bcast_sockets[bcast.toString()] = udp_sock
                    udp_sock.readyRead.connect(self.read_ip_from_socket)

    def get_ip_button_clicked(self):
        for ip, sock in self.bcast_sockets.items():
            sock.writeDatagram("upms_1v_get_ip".encode('ascii'), QtNetwork.QHostAddress(ip), 5007)

    @exception_decorator
    def read_ip_from_socket(self):
        self.ui.ip_combobox.clear()
        for ip, sock in self.bcast_sockets.items():
            while sock.hasPendingDatagrams():
                datagram = sock.receiveDatagram()
                data = bytes(datagram.data()).decode(encoding='ascii')
                ip, _ = data.split(';')
                try:
                    socket.inet_aton(ip)
                    self.ui.ip_combobox.addItem(ip)
                    print(ip)
                except socket.error:
                    logging.error(f"{ip} не является валидным ip адресом")

    @exception_decorator_print
    def input_result_button_clicked(self, _):
        rows = self.ui.measures_table.selectionModel().selectedRows()
        if rows:
            for idx in rows:
                real_row = self.proxy.mapToSource(idx).row()

                self.ui.measures_table.selectionModel().setCurrentIndex(
                    self.measures_table_model.index(idx.row(), 0),
                    QtCore.QItemSelectionModel.ClearAndSelect | QtCore.QItemSelectionModel.Rows
                )
                files_path = self.ui.download_path_edit.text()
                if files_path and os.path.isdir(files_path):
                    upms_measure = self.measures_table_model.get_upms_measure_by_row(real_row)
                    dialog = ResultInputDialog(upms_measure, files_path, self.settings, self)
                    res = dialog.exec()
                    if res == QtWidgets.QDialog.Accepted:
                        new_result = dialog.get_result()
                        self.measures_table_model.update_result(real_row, new_result)
                        logging.warning(upms_measure.result)
                        # Для сохранения состояния (иначе не вызывается closeEvent)
                        dialog.close()
                    else:
                        break
                else:
                    QtWidgets.QMessageBox.critical(self, Text.get("err"), Text.get("path_err"),
                                                   QtWidgets.QMessageBox.Ok, QtWidgets.QMessageBox.Ok)
        else:
            QtWidgets.QMessageBox.information(self, Text.get("info"), Text.get("selection_info"),
                                              QtWidgets.QMessageBox.Ok, QtWidgets.QMessageBox.Ok)

    def get_template_file(self, a_measure_type) -> str:
        for path in self.MEASURE_TYPE_TO_TEMPLATE_PATH[a_measure_type]:
            if os.path.isfile(path):
                return path
        return ""

    @staticmethod
    def get_accessible_name(a_name_template, a_save_folder, a_extension: str) -> str:
        number = 2
        name = a_name_template
        while os.path.isfile(f"{a_save_folder.rstrip(os.sep)}{os.sep}{name}{a_extension}"):
            name = f"{a_name_template}_{number}"
            number += 1
        return f"{name}{a_extension}"

    @staticmethod
    def insert_upms_measures_into_sheet(a_workbook: openpyxl.Workbook, a_sheet_name,
                                        a_upms_measures: List[UpmsMeasure]):
        sheet = a_workbook.get_sheet_by_name(a_sheet_name)
        for idx, column in enumerate(sheet.iter_cols(min_col=2, max_col=len(a_upms_measures)+1, min_row=1, max_row=5)):
            column[0].value = a_upms_measures[idx].id
            column[1].value = a_upms_measures[idx].date
            column[2].value = a_upms_measures[idx].interval
            column[3].value = a_upms_measures[idx].result
            column[4].value = a_upms_measures[idx].comment

    @staticmethod
    def insert_photos_into_sheet(a_photos_path: str, a_workbook: openpyxl.Workbook, a_sheet_name,
                                 a_upms_measures: List[UpmsMeasure]):
        if a_sheet_name in a_workbook.sheetnames:
            sheet = a_workbook.get_sheet_by_name(a_sheet_name)
        else:
            sheet = a_workbook.create_sheet(a_sheet_name)

        vertical_start = 3
        vertical_step = 27

        horizontal_start = 2
        horizontal_step = 9
        for idx, upms_measure in enumerate(a_upms_measures):
            photo_path = a_photos_path.rstrip(os.sep) + os.sep + f"{upms_measure.id}.jpg"
            try:
                img = openpyxl.drawing.image.Image(photo_path)
                new_height = img.height * 2 / 3
                new_width = img.width * 2 / 3
                img.height = new_height
                img.width = new_width

                row = vertical_start + idx // 2 * vertical_step if idx % 2 == 0 else \
                    vertical_start + idx // 2 * vertical_step + 1
                col = horizontal_start if idx % 2 == 0 else \
                    (horizontal_start + horizontal_step)

                sheet.add_image(img, sheet.cell(row, col).coordinate)
                cell_font = openpyxl.styles.Font(size='15')
                sheet.cell(row - 1, col, value=Text.get("measure").format(upms_measure.id)).font = cell_font
            except FileNotFoundError:
                logging.warning(f"File {photo_path} is not found")

    @staticmethod
    def insert_extra_parameters_into_sheet(a_workbook: openpyxl.Workbook, a_sheet_name,
                                           a_extra_parameters: List[Tuple[str, str]]):
        sheet = a_workbook.get_sheet_by_name(a_sheet_name)
        row = 8
        for parameter, value in a_extra_parameters:
            sheet.cell(row, 1).value = parameter
            sheet.cell(row, 2).value = value
            sheet.row_dimensions[row].height = 30
            row += 1

    @exception_decorator
    def create_report(self, a_name_template, a_save_folder, a_template_path, a_photos_path,
                      a_upms_measures: List[UpmsMeasure]):
        filename = self.get_accessible_name(a_name_template, a_save_folder, os.path.splitext(a_template_path)[1])
        report_path = a_save_folder.rstrip(os.sep) + os.sep + filename
        shutil.copyfile(a_template_path, report_path)

        data_sheet = ""
        photo_sheet = ""
        data_sheet_exists = True
        wb = openpyxl.load_workbook(report_path)
        if self.data_sheet_ru in wb.sheetnames:
            data_sheet = self.data_sheet_ru
            photo_sheet = self.photo_sheet_ru
        elif self.data_sheet_en in wb.sheetnames:
            data_sheet = self.data_sheet_en
            photo_sheet = self.photo_sheet_en
        else:
            data_sheet_exists = False

        if data_sheet_exists:
            self.insert_upms_measures_into_sheet(wb, data_sheet, a_upms_measures)
            self.insert_extra_parameters_into_sheet(wb, data_sheet, self.db.get_parameters())
            self.insert_photos_into_sheet(a_photos_path, wb, photo_sheet, a_upms_measures)

            wb.save(report_path)
            wb.close()

            QtWidgets.QMessageBox.information(self, Text.get("info"), Text.get("success_generated").format(report_path),
                                              QtWidgets.QMessageBox.Ok, QtWidgets.QMessageBox.Ok)
            os.startfile(report_path)
        else:
            QtWidgets.QMessageBox.critical(self, Text.get("err"), Text.get("data_sheet_not_found").format(report_path),
                                           QtWidgets.QMessageBox.Ok, QtWidgets.QMessageBox.Ok)
            wb.close()

    def create_report_button_clicked(self, _):
        rows = self.ui.measures_table.selectionModel().selectedRows()
        if rows:
            types = [self.measures_table_model.get_type(self.proxy.mapToSource(idx).row()) for idx in rows]
            if len(types) == types.count(types[0]):
                save_folder = self.ui.save_folder_edit.text()
                if save_folder and os.path.isdir(save_folder):
                    template_path = self.get_template_file(types[0])
                    if template_path:
                        if self.ui.download_path_edit.text():
                            name_template = self.ui.name_template_edit.text() if self.ui.name_template_edit.text() else \
                                self.default_name_template
                            upms_measures = [self.measures_table_model.get_upms_measure_by_row(self.proxy.mapToSource(idx).row())
                                             for idx in rows]
                            self.create_report(name_template, save_folder, template_path,
                                               self.ui.download_path_edit.text(), upms_measures)
                        else:
                            QtWidgets.QMessageBox.critical(self, Text.get("err"), Text.get("path_err"),
                                                           QtWidgets.QMessageBox.Ok, QtWidgets.QMessageBox.Ok)
                    else:
                        QtWidgets.QMessageBox.critical(self, Text.get("err"), Text.get("templates_are_not_found"),
                                                       QtWidgets.QMessageBox.Ok, QtWidgets.QMessageBox.Ok)
                else:
                    QtWidgets.QMessageBox.critical(self, Text.get("err"), Text.get("save_folder_error"),
                                                   QtWidgets.QMessageBox.Ok, QtWidgets.QMessageBox.Ok)
            else:
                QtWidgets.QMessageBox.critical(self, Text.get("err"), Text.get("same_type_err"),
                                               QtWidgets.QMessageBox.Ok, QtWidgets.QMessageBox.Ok)
        else:
            QtWidgets.QMessageBox.information(self, Text.get("info"), Text.get("selection_info"),
                                              QtWidgets.QMessageBox.Ok, QtWidgets.QMessageBox.Ok)

    def remove_selected_button_clicked(self, _):
        rows = self.ui.measures_table.selectionModel().selectedRows()
        if rows:
            res = QtWidgets.QMessageBox.question(self, Text.get("confirm"), Text.get("delete_confirm"),
                                                 QtWidgets.QMessageBox.Ok | QtWidgets.QMessageBox.Cancel,
                                                 QtWidgets.QMessageBox.Cancel)
            if res == QtWidgets.QMessageBox.Ok:
                real_rows = [self.proxy.mapToSource(r).row() for r in rows]
                real_rows.sort(reverse=True)
                for row in real_rows:
                    self.measures_table_model.remove_row(row)
        else:
            QtWidgets.QMessageBox.information(self, Text.get("info"), Text.get("selection_info"),
                                              QtWidgets.QMessageBox.Ok, QtWidgets.QMessageBox.Ok)

    def extra_params_button_clicked(self, _):
        old_parameters = self.db.get_parameters()
        dialog = ExtraParametersDialog(old_parameters, self.settings)
        res = dialog.exec()
        if res == QtWidgets.QDialog.Accepted:
            new_parameters = dialog.get_parameters()
            self.db.set_parameters(new_parameters)
        dialog.close()

    def choose_download_path_button_clicked(self):
        new_path = QtWidgets.QFileDialog.getExistingDirectory(self, Text.get("download_folder"), self.settings.path)
        if new_path:
            self.ui.download_path_edit.setText(new_path)

    def choose_save_folder_path_button_clicked(self):
        new_path = QtWidgets.QFileDialog.getExistingDirectory(self, Text.get("save_folder"), self.settings.save_folder)
        if new_path:
            self.ui.save_folder_edit.setText(new_path)

    def russian_language_chosen(self):
        self.install_translator(Text.Lang.RU)
        self.settings.language = int(Text.Lang.RU)

    def english_language_chosen(self):
        self.install_translator(Text.Lang.EN)
        self.settings.language = int(Text.Lang.EN)

    def open_about(self):
        about_dialog = AboutDialog(self)
        about_dialog.exec()

    def closeEvent(self, a_event: QtGui.QCloseEvent):
        self.settings.save_qwidget_state(self.ui.measures_table)
        self.settings.save_qwidget_state(self.ui.mw_splitter_1)
        self.settings.save_qwidget_state(self)
        a_event.accept()
